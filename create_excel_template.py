"""
Builds Invoice_Data.xlsx - the phase 2 data source for generate_invoice.py.

Sheet layout: two columns. Column A is the field label, column B is the
value. A blank row marks a section break - that's purely visual, to make the
sheet easier to read; generate_invoice.py's loader matches rows by their
label text, not by position, so section breaks can move or be skipped
without breaking anything.

To edit the invoice data: open Invoice_Data.xlsx directly and change the
values in column B. Re-run this script only if you want to regenerate the
template from scratch with different sample defaults, or a different
signature.

Signature: the Signature row's value cell holds a filename (e.g.
"signature_sample.png") - generate_invoice.py looks for that file next to
Invoice_Data.xlsx and inserts it as a picture. This is the recommended way
to hand off the invoice generator to someone else: send generate_invoice.py,
Invoice_Data.xlsx, and the signature PNG together, and it just works
regardless of Excel version or how the files were copied. (Embedding a
picture directly in the sheet also still works as a fallback if the
Signature cell is left blank - see load_invoice_data_from_excel().)

Do not rename the labels in column A - generate_invoice.py looks them up by
this exact text (see EXCEL_FIELD_LABELS in that file).
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from generate_invoice import EXCEL_FIELD_LABELS as L

HERE = os.path.dirname(os.path.abspath(__file__))


def build_template(out_path=None, signature_filename="signature_sample.png"):
    out_path = out_path or os.path.join(HERE, "Invoice_Data.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Data"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45

    # None entries are blank-row section breaks (see module docstring).
    rows = [
        (L["invoice_no"], "INV-214"),
        (L["invoice_date"], "24-Jun-2026"),
        (L["due_date"], "26-Jun-2026"),
        (L["place_of_supply"], "36-Telangana"),
        None,
        (L["bill_to_name"], "HEARTCULTURE NATURAL PRODUCT LLP"),
        (L["bill_to_address"], "13-110 Kanha village nandigama mandal\nRanga Reddy telangana 509325"),
        (L["bill_to_state"], "36-Telangana"),
        (L["bill_to_gstin"], "36AAMFH4243C1ZL"),
        None,
        (L["shipping_address"], "Rajasthan"),
        None,
        (L["transport_name"], "-"),
        (L["transport_delivery_date"], "-"),
        (L["transport_vehicle_number"], "RJ 19 GF 8254"),
        (L["transport_delivery_location"], "Rajasthan"),
        None,
        (L["item_name"], "Transport charges for Wakavali to khana hyderabad"),
        (L["item_quantity"], 10),
        (L["item_unit"], "Ton"),
        (L["item_price_per_unit"], 4900.00),
        None,
        (L["amount_in_words"], "Fifty One Thousand Four Hundred and Fifty Rupees only"),
        None,
        (L["bank_name"], "Hdfc Bank, Serilingampally"),
        (L["bank_account_no"], "50200118563217"),
        (L["bank_ifsc"], "HDFC0002073"),
        (L["bank_account_holder"], "SLNS LOGISTICS"),
        None,
        (L["signature"], signature_filename),
    ]

    label_font = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")

    row_idx = 1
    for entry in rows:
        if entry is None:
            row_idx += 1
            continue
        label, value = entry
        ws.cell(row=row_idx, column=1, value=label).font = label_font
        if value is not None:
            cell = ws.cell(row=row_idx, column=2, value=value)
            if isinstance(value, str) and "\n" in value:
                cell.alignment = wrap_top
        row_idx += 1

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    path = build_template()
    print(f"Excel template generated: {path}")
